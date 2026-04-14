import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df_raw = pd.read_csv(
    'C:/Users/Admin/hipython/ml/bermuda.project/encar_kia_all/output/kia_detail_batch.csv',
    encoding='utf-8'
)

META = [
    ("매물ID",         "매물 고유 식별자",         "ID",   "기존선택", "사용안함(ID)",             "int",      "학습 제외"),
    ("제조사",         "차량 제조사",              "입력", "제거권장", "제거(단일값:기아)",         "—",        "전체 기아 단일값"),
    ("모델",           "차량 모델명",              "입력", "기존선택", "Target Encoding",          "float",    "모델별 평균가로 인코딩"),
    ("현재가격_만원",   "현재 매물 가격(만원)",     "타겟", "기존선택", "타겟 변수",                "int",      "Y: 예측 대상"),
    ("연식",           "차량 연식(년도)",          "입력", "기존선택", "파생변수 생성",            "int",      "차량연령=2026-연식"),
    ("주행거리_km",    "주행거리(km)",             "입력", "기존선택", "정수형 그대로 사용",       "int",      "이상치 처리 필요"),
    ("연료",           "연료 유형",               "입력", "기존선택", "One-Hot Encoding",         "int",      "6개 범주"),
    ("색상",           "차량 색상",               "입력", "기존선택", "희소통합+Label Encoding",  "int",      "10건 미만 -> 기타"),
    ("지역",           "등록 지역",               "입력", "기존선택", "One-Hot Encoding",         "int",      "15개 지역"),
    ("변속기",         "변속기 종류",             "입력", "기존선택", "Binary Enc.(오토=1)",      "int",      "오토/수동 2개 범주"),
    ("배기량_cc",      "배기량(cc)",              "입력", "기존선택", "정수형 그대로 사용",       "int",      "전기차=0 처리 필요"),
    ("차급",           "차량 등급(차급)",         "입력", "기존선택", "One-Hot Encoding",         "int",      "9개 범주"),
    ("좌석수",         "좌석 수",                 "입력", "기존선택", "정수형 그대로 사용",       "int",      "결측 0.1%"),
    ("상사명",         "딜러 상사명",             "입력", "기존선택", "Target Encoding",          "float",    "1007개 고카디널리티"),
    ("압류건수",       "압류 건수",               "입력", "기존선택", "정수형 그대로 사용",       "int",      "결측 3.8%"),
    ("저당건수",       "저당 건수",               "입력", "기존선택", "정수형 그대로 사용",       "int",      "결측 3.8%"),
    ("보험이력공개여부","보험이력 공개 여부/건수", "입력", "기존선택", "건수 수치 추출",           "int",      "비공개->NaN, 공개->-1"),
    ("보험이력건수",   "보험이력 건수",           "입력", "기존선택", "정수형 그대로 사용",       "int",      "결측 8.9%"),
    ("성능점검유형",   "성능점검 유형",           "입력", "기존선택", "Binary Enc.(엔카직영=1)",  "int",      "2개 범주"),
    ("옵션_기본코드개수","기본 옵션 개수",         "입력", "기존선택", "정수형 그대로 사용",       "int",      ""),
    ("옵션_선택코드개수","선택 옵션 개수",         "입력", "기존선택", "정수형 그대로 사용",       "int",      ""),
    ("옵션_튜닝코드개수","튜닝 옵션 개수",         "입력", "기존선택", "정수형 그대로 사용",       "int",      ""),
    ("옵션_기타개수",  "기타 옵션 개수",          "입력", "기존선택", "정수형 그대로 사용",       "int",      ""),
    ("주요옵션_선루프","주요옵션: 선루프",        "입력", "기존선택", "Boolean 그대로 사용",      "boolean",  "결측 0.6%"),
    ("주요옵션_헤드램프 (LED)","주요옵션: LED 헤드램프","입력","기존선택","Boolean 그대로 사용","boolean","결측 39.6%->0으로 대체"),
    ("주요옵션_주차감지센서","주요옵션: 주차감지센서","입력","기존선택","Boolean 그대로 사용","boolean","결측 0.6%"),
    ("주요옵션_후방카메라","주요옵션: 후방카메라","입력","기존선택","Boolean 그대로 사용",   "boolean","결측 0.6%"),
    ("주요옵션_자동에어컨","주요옵션: 자동에어컨","입력","기존선택","Boolean 그대로 사용",   "boolean","결측 0.6%"),
    ("주요옵션_스마트키","주요옵션: 스마트키",   "입력","기존선택","Boolean 그대로 사용",   "boolean","결측 0.6%"),
    ("주요옵션_내비게이션","주요옵션: 내비게이션","입력","기존선택","Boolean 그대로 사용",  "boolean","결측 0.6%"),
    ("주요옵션_열선시트","주요옵션: 열선시트",   "입력","기존선택","Boolean 그대로 사용",   "boolean","결측 0.6%"),
    ("주요옵션_통풍시트","주요옵션: 통풍시트",   "입력","기존선택","Boolean 그대로 사용",   "boolean","결측 0.6%"),
    ("주요옵션_가죽시트","주요옵션: 가죽시트",   "입력","기존선택","Boolean 그대로 사용",   "boolean","결측 0.6%"),
    ("교환_개수",      "부품 교환 개수",          "입력","기존선택","정수형 그대로 사용",    "int",     "결측 7.7%"),
    ("판금_개수",      "판금 개수",               "입력","기존선택","정수형 그대로 사용",    "int",     "결측 7.7%"),
    ("부식_여부",      "부식 여부",               "입력","기존선택","Boolean 그대로 사용",   "boolean","결측 7.7%"),
    ("교환_여부",      "교환 이력 여부",          "입력","기존선택","Boolean 그대로 사용",   "boolean","결측 7.7%"),
    ("판금_여부",      "판금 이력 여부",          "입력","기존선택","Boolean 그대로 사용",   "boolean","결측 7.7%"),
    ("외판수리_여부",  "외판 수리 여부",          "입력","기존선택","Boolean 그대로 사용",   "boolean","결측 7.7%"),
    ("보험이력_여부",  "보험이력 존재 여부",      "입력","기존선택","Boolean 그대로 사용",   "boolean","결측 8.9%"),
    ("사고종합_여부",  "종합 사고 여부",          "입력","기존선택","Boolean 그대로 사용",   "boolean","결측 2.6%"),
    # 추가권장
    ("등급명",         "차량 등급명(엔진/구동)",  "입력","추가권장","Target Encoding",       "float",   "309개 범주, 가격 영향 큼"),
    ("세부트림",       "세부 트림명",             "입력","추가권장","Target Encoding",       "float",   "결측 48.5%->없음 처리"),
    ("엔카진단여부값", "엔카 진단 수준(-1~2)",    "입력","추가권장","Ordinal 정수 그대로",   "int",     "신뢰도 지표: -1,0,1,2"),
    ("사고강도점수",   "사고 심각도 점수",        "입력","추가권장","정수형 그대로 사용",    "float",   "결측 7.7%, 범위 0~24"),
    ("중대사고_추정",  "중대사고 추정 여부",      "입력","추가권장","Boolean 그대로 사용",   "boolean","결측 7.7%"),
    ("부식_위험",      "부식 위험 여부",          "입력","추가권장","Boolean 그대로 사용",   "boolean","결측 7.7%"),
    ("조회수",         "매물 조회수",             "입력","추가권장","로그 변환 후 사용",     "float",   "수요 프록시 변수"),
    ("찜수",           "매물 찜 수",              "입력","추가권장","로그 변환 후 사용",     "float",   "수요 프록시 변수"),
    ("판매자유형",     "판매자 유형(매매/개인)",  "입력","추가권장","One-Hot Encoding",      "int",     "결측 47.7%->미상 처리"),
    # 제거권장
    ("차량번호",       "차량 번호판",             "메타","제거권장","제거(식별자+개인정보)",  "—",       ""),
    ("VIN",            "차대번호",               "메타","제거권장","제거(식별자)",          "—",       ""),
    ("연식_원문",      "연식 원문 텍스트",        "메타","제거권장","제거(연식 컬럼으로 대체)","—",      ""),
    ("주행거리_원문",  "주행거리 원문 텍스트",    "메타","제거권장","제거(주행거리_km 대체)", "—",       ""),
    ("차량명",         "차량 전체명",             "메타","제거권장","제거(모델+등급명 중복)", "—",       ""),
    ("등록번호",       "매물 등록번호",           "메타","제거권장","제거(식별자)",          "—",       "매물ID와 동일"),
    ("해시태그",       "해시태그",               "메타","제거권장","제거(결측 65.7%+NLP)",  "—",       ""),
    ("실촬영문구",     "실촬영 문구",             "메타","제거권장","제거(결측 52.7%+NLP)",  "—",       ""),
    ("딜러명",         "딜러 이름",              "메타","제거권장","제거(상사명 대체)",      "—",       ""),
    ("판매중대수",     "딜러 현재 판매 대수",     "메타","제거권장","제거(결측 47.8%)",      "—",       ""),
    ("판매완료대수",   "딜러 판매 완료 대수",     "메타","제거권장","제거(결측 47.8%)",      "—",       ""),
    ("판매자지역",     "판매자 상세 주소",        "메타","제거권장","제거(지역 컬럼 대체)",  "—",       ""),
    ("종사원증번호",   "딜러 종사원증 번호",      "메타","제거권장","제거(식별자)",          "—",       ""),
    ("진단센터명",     "진단센터 이름",           "메타","제거권장","제거(엔카진단여부값 대체)","—",     ""),
    ("진단센터전화",   "진단센터 전화번호",       "메타","제거권장","제거(불필요)",          "—",       ""),
    ("진단센터주소",   "진단센터 주소",           "메타","제거권장","제거(불필요)",          "—",       ""),
    ("엔카믿고값",     "엔카 믿고 여부(-1/0)",    "메타","제거권장","제거(유효값 없음)",     "—",       "값이 -1,0만 존재"),
    ("광고한줄문구",   "광고 문구",              "메타","제거권장","제거(결측 59.3%+NLP)",  "—",       ""),
    ("성능점검설명",   "성능점검 상세 설명",      "메타","제거권장","제거(결측 93.3%)",      "—",       ""),
    ("requestUrl",     "API 요청 URL",           "메타","제거권장","제거(크롤링 메타)",     "—",       ""),
    ("pageAccessToken","API 액세스 토큰",         "메타","제거권장","제거(크롤링 메타)",     "—",       ""),
    ("옵션_기본코드",  "기본 옵션 코드 원문",     "메타","제거권장","제거(개수 컬럼 대체)",  "—",       ""),
    ("옵션_선택코드",  "선택 옵션 코드 원문",     "메타","제거권장","제거(개수 컬럼 대체)",  "—",       ""),
    ("옵션_튜닝코드",  "튜닝 옵션 코드 원문",     "메타","제거권장","제거(개수 컬럼 대체)",  "—",       ""),
    ("옵션_기타값",    "기타 옵션 값 원문",       "메타","제거권장","제거(결측 93.6%+NLP)",  "—",       ""),
    ("성능기록부_raw", "성능기록부 원문",         "메타","제거권장","제거(파생변수 존재)",   "—",       "교환/판금/부식으로 대체"),
    ("상세링크",       "매물 상세 URL",           "메타","제거권장","제거(URL)",             "—",       ""),
    ("주요옵션_헤드램프 (HID)","주요옵션: HID 헤드램프","입력","제거권장","제거(결측 61.1%)", "—",      "LED 헤드램프로 대체"),
    ("수집시각",       "데이터 수집 시각",        "메타","제거권장","제거(크롤링 메타)",     "—",       ""),
]

stats = {}
for col in df_raw.columns:
    null_pct = round(df_raw[col].isna().mean() * 100, 1)
    n_unique = df_raw[col].nunique()
    raw_dtype = str(df_raw[col].dtype)
    non_null = df_raw[col].dropna()
    if raw_dtype in ['int64', 'float64'] and len(non_null) > 0:
        rng = f"{round(non_null.min(),1)} ~ {round(non_null.max(),1)}"
    else:
        vals = non_null.head(2).tolist()
        rng = ' / '.join(str(v)[:15] for v in vals)
    stats[col] = {'raw_dtype': raw_dtype, 'null_pct': null_pct,
                  'n_unique': n_unique, 'range_example': rng}

wb = Workbook()

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

COLORS = {
    "기존선택": "D9EAD3",
    "추가권장": "FFF2CC",
    "제거권장": "FCE4D6",
}

# ── Sheet 1: 전체 피처 정의서 ──────────────────────────────────────
ws1 = wb.active
ws1.title = "전체 피처 정의서"

ws1.merge_cells("A1:L1")
t1 = ws1["A1"]
t1.value = "기아 중고차 가격예측 프로젝트 — 전체 피처 정의서"
t1.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
t1.fill = PatternFill("solid", fgColor="1F4E79")
t1.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

headers1 = ["No.", "CSV 컬럼명", "한글 설명", "원본 데이터타입",
            "결측률(%)", "고유값수", "값 범위/예시",
            "피처 역할", "선택 여부", "처리 방법", "결과 데이터타입", "비고"]
col_widths1 = [5, 26, 22, 14, 10, 8, 30, 8, 10, 28, 14, 28]

for ci, (h, w) in enumerate(zip(headers1, col_widths1), 1):
    cell = ws1.cell(row=2, column=ci, value=h)
    cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = center_align
    ws1.column_dimensions[get_column_letter(ci)].width = w
ws1.row_dimensions[2].height = 30

meta_dict = {row[0]: row for row in META}
all_col_order = list(df_raw.columns)
no_counter = {"기존선택": 0, "추가권장": 0, "제거권장": 0}

row_idx = 3
for col_name in all_col_order:
    s = stats.get(col_name, {})
    if col_name in meta_dict:
        m = meta_dict[col_name]
        col_nm, kor, role, sel, proc, res_type, note = m
    else:
        col_nm, kor, role, sel, proc, res_type, note = col_name, "—", "메타", "제거권장", "제거", "—", ""

    no_counter[sel] = no_counter.get(sel, 0) + 1
    rf = PatternFill("solid", fgColor=COLORS.get(sel, "FFFFFF"))

    row_data = [no_counter[sel], col_name, kor,
                s.get("raw_dtype",""), s.get("null_pct",""), s.get("n_unique",""),
                s.get("range_example",""), role, sel, proc, res_type, note]

    for ci, val in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=ci, value=val)
        cell.fill = rf
        cell.border = border
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(
            horizontal="center" if ci in [1,4,5,6,8,9,11] else "left",
            vertical="center", wrap_text=True)
    ws1.row_dimensions[row_idx].height = 18
    row_idx += 1

ws1.freeze_panes = "A3"
ws1.auto_filter.ref = f"A2:L{row_idx-1}"

# ── Sheet 2: 최종 선택 피처 요약 ─────────────────────────────────
ws2 = wb.create_sheet("최종 선택 피처 요약")

ws2.merge_cells("A1:J1")
t2 = ws2["A1"]
t2.value = "기아 중고차 가격예측 — 최종 선택 피처 요약"
t2.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
t2.fill = PatternFill("solid", fgColor="375623")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

h2_headers = ["No.", "CSV 컬럼명", "한글 설명", "피처 역할",
              "원본 타입", "결측률(%)", "고유값수", "처리 방법", "결과 타입", "비고"]
h2_widths = [5, 26, 22, 8, 10, 10, 8, 28, 10, 28]
for ci, w in enumerate(h2_widths, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

GROUPS = [
    ("◆ 타겟 변수 (1개)",
     [m for m in META if m[0] == "현재가격_만원"],
     "C0504D", "FFF0F0"),
    ("◆ 기존 선택 피처 (41개: ID 포함)",
     [m for m in META if m[3] == "기존선택"],
     "1F4E79", "D9EAD3"),
    ("◆ 추가 권장 피처 (9개)",
     [m for m in META if m[3] == "추가권장"],
     "7B5A2B", "FFF2CC"),
]

cur_row = 2
for grp_label, rows, grp_color, row_color in GROUPS:
    ws2.merge_cells(f"A{cur_row}:J{cur_row}")
    gc = ws2.cell(row=cur_row, column=1, value=grp_label)
    gc.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    gc.fill = PatternFill("solid", fgColor=grp_color)
    gc.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[cur_row].height = 22
    cur_row += 1

    for ci, h in enumerate(h2_headers, 1):
        cell = ws2.cell(row=cur_row, column=ci, value=h)
        cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = center_align
        cell.border = border
    ws2.row_dimensions[cur_row].height = 22
    cur_row += 1

    for no_i, m in enumerate(rows, 1):
        col_nm, kor, role, sel, proc, res_type, note = m
        s = stats.get(col_nm, {})
        rf = PatternFill("solid", fgColor=row_color)
        row_vals = [no_i, col_nm, kor, role,
                    s.get("raw_dtype",""), s.get("null_pct",""),
                    s.get("n_unique",""), proc, res_type, note]
        for ci, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=cur_row, column=ci, value=val)
            cell.fill = rf
            cell.border = border
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(
                horizontal="center" if ci in [1,4,5,6,7,9] else "left",
                vertical="center", wrap_text=True)
        ws2.row_dimensions[cur_row].height = 18
        cur_row += 1
    cur_row += 1

ws2.freeze_panes = "A3"

# ── Sheet 3: 인코딩 전략 상세 ────────────────────────────────────
ws3 = wb.create_sheet("인코딩 전략 상세")

ws3.merge_cells("A1:G1")
t3 = ws3["A1"]
t3.value = "str 타입 변수 인코딩 전략 상세"
t3.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
t3.fill = PatternFill("solid", fgColor="7030A0")
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

enc_headers = ["No.", "변수명", "고유값수", "인코딩 방법", "선택 이유", "주의사항", "결과 컬럼 수"]
enc_widths   = [5, 22, 8, 22, 40, 40, 12]
for ci, (h, w) in enumerate(zip(enc_headers, enc_widths), 1):
    cell = ws3.cell(row=2, column=ci, value=h)
    cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="7030A0")
    cell.alignment = center_align
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.row_dimensions[2].height = 28

ENC_DATA = [
    (1,"모델",102,"Target Encoding",
     "고카디널리티(102개). OHE시 차원 폭발. 모델별 평균가로 가격 정보 압축",
     "반드시 train 기준으로만 계산 -> data leakage 방지",1),
    (2,"연료",6,"One-Hot Encoding",
     "범주 수 적음(6개). 순서 없는 명목형. 희소성 없음",
     "전기차(배기량=0) 처리와 연동 확인",6),
    (3,"색상",24,"희소통합 + Label Encoding",
     "24개 중 10건 미만 희소값을 기타로 묶어 약 15개로 축소 후 Label Encoding",
     "색상은 순서 없음. OHE도 고려 가능",1),
    (4,"지역",15,"One-Hot Encoding",
     "범주 수 적당(15개). 순서 없는 명목형",
     "경기 쏠림(45.7%) 주의",15),
    (5,"변속기",2,"Binary Encoding",
     "2개 범주(오토/수동). Binary로 충분",
     "수동 데이터 극소(7건) -> 과적합 주의",1),
    (6,"차급",9,"One-Hot Encoding",
     "범주 수 적당(9개). 순서 없는 명목형",
     "화물차/스포츠카 극소 -> 기타 통합 고려",9),
    (7,"상사명",1007,"Target Encoding",
     "고카디널리티(1007개). 딜러 신뢰도/규모가 가격에 영향",
     "train 기준 계산 필수. 신규 상사명은 전체 평균 대체",1),
    (8,"보험이력공개여부",21,"수치 추출 변환",
     "보험이력 N건에서 정규식으로 N 추출해 숫자형으로. 비공개->NaN, 공개->-1",
     "Feature 목록의 boolean 결과타입과 불일치 -> 재협의 필요",1),
    (9,"성능점검유형",2,"Binary Encoding",
     "2개 범주(일반/엔카직영). 엔카직영이 더 신뢰도 높아 가격에 영향",
     "결측 2.2% -> 최빈값(일반)으로 대체",1),
    (10,"등급명(추가권장)",309,"Target Encoding",
     "고카디널리티(309개). 등급별 평균가로 압축. 가격 설명력 높음",
     "모델과 등급명은 높은 상관관계 -> 다중공선성 주의",1),
    (11,"세부트림(추가권장)",81,"Target Encoding",
     "결측 48.5% -> 없음 카테고리로 대체 후 인코딩",
     "결측을 없음으로 처리시 정보 손실 없음",1),
    (12,"판매자유형(추가권장)",3,"One-Hot Encoding",
     "3개 범주. 결측 47.7% -> 미상 카테고리 추가",
     "결측 비율 높아 미상 비율 큰 피처",4),
]

for row_data in ENC_DATA:
    r_idx = row_data[0] + 2
    fill_color = "EAD1DC" if row_data[0] % 2 == 0 else "F3E6FF"
    rf = PatternFill("solid", fgColor=fill_color)
    for ci, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r_idx, column=ci, value=val)
        cell.fill = rf
        cell.border = border
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(
            horizontal="center" if ci in [1,3,7] else "left",
            vertical="center", wrap_text=True)
    ws3.row_dimensions[r_idx].height = 45

ws3.freeze_panes = "A3"

out_path = "C:/Users/Admin/hipython/ml/bermuda.project/데이터_정의서_KIA중고차가격예측.xlsx"
wb.save(out_path)
print(f"저장 완료: {out_path}")
